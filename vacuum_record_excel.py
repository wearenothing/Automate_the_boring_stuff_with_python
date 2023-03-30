def generate_vacuum_record():
    """
    Generate vacuum recording excel, return filename
    """
    # !/usr/bin/env python
    # coding: utf-8

    import openpyxl
    import datetime
    from openpyxl.styles import Font, Color, Alignment, Border, Side
    from openpyxl.utils import get_column_letter, column_index_from_string

    wb = openpyxl.Workbook()

    sheet0 = wb.active
    sheet0.title = '真空运行'
    sheet1 = wb.create_sheet('烘烤温度')

    # In[125]:

    sheet0.merge_cells('A1:Q2')
    sheet0['A1'] = "2023年春季EAST真空运行实验数据记录表"

    # In[126]:

    sheet0['A3'] = '日期'
    sheet0['B3'] = (datetime.date.today() - datetime.timedelta(days=1)).isoformat()  # time.strftime("%Y%m%d")
    sheet0['C3'] = '值班人员'
    sheet0.merge_cells('D3:Q3')
    # sheet0['D3'] = 'People'

    # In[127]:

    sheet0.merge_cells('A4:A5')
    sheet0.merge_cells('B4:C4')
    sheet0.merge_cells('D4:E4')
    sheet0.merge_cells('N4:P4')
    sheet0.merge_cells('Q4:Q5')
    sheet0['A4'] = '时间'
    sheet0['B4'] = '内真空'
    sheet0['D4'] = '外真空'
    sheet0['F4'] = '传输线'
    sheet0['G4'] = '5对引线罐'
    sheet0['H4'] = '5对传输线'
    sheet0['I4'] = '8对引线罐'
    sheet0['J4'] = '8对传输线'
    sheet0['K4'] = '阀箱'
    sheet0['L4'] = '新阀箱'
    sheet0['M4'] = '低温传输线'
    sheet0['N4'] = '电子回旋'
    sheet0['Q4'] = '记录人'

    # In[128]:

    sheet0['B5'] = 'G1.1'
    sheet0['C5'] = 'G1.7'
    sheet0['D5'] = 'G2.3'
    sheet0['E5'] = 'G2.4'
    sheet0['F5'] = 'G5.1'
    sheet0['G5'] = 'G5.4'
    sheet0['H5'] = 'G5.7'
    sheet0['I5'] = 'G5.3'
    sheet0['J5'] = 'G5.8'
    sheet0['K5'] = 'G5.5'
    sheet0['L5'] = 'G5.2'
    sheet0['M5'] = 'G5.6'
    sheet0['N5'] = 'GE2.1'
    sheet0['O5'] = 'GE3.1'
    sheet0['P5'] = 'GE4.1'

    # In[129]:

    for i in range(6, 18):
        sheet0[f'A{i}'] = datetime.time(hour=i + 2).strftime("%H:%M")

    for i in range(18, 31):
        sheet0.merge_cells(f'A{i}:Q{i}')

    sheet0.merge_cells('A31:Q32')
    sheet0['A31'] = '=A1'

    sheet0['A33'] = '日期'
    sheet0['B33'] = '=B3'  # time.isoformat
    sheet0['C33'] = '值班人员'
    sheet0.merge_cells('D33:Q33')
    sheet0['D33'] = '=D3'

    sheet0.merge_cells('A34:A35')
    sheet0.merge_cells('B34:C34')
    sheet0.merge_cells('D34:E34')
    sheet0.merge_cells('N34:P34')
    sheet0.merge_cells('Q34:Q35')
    sheet0['A34'] = '时间'
    sheet0['B34'] = '内真空'
    sheet0['D34'] = '外真空'
    sheet0['F34'] = '传输线'
    sheet0['G34'] = '5对引线罐'
    sheet0['H34'] = '5对传输线'
    sheet0['I34'] = '8对引线罐'
    sheet0['J34'] = '8对传输线'
    sheet0['K34'] = '阀箱'
    sheet0['L34'] = '新阀箱'
    sheet0['M34'] = '低温传输线'
    sheet0['N34'] = '电子回旋'
    sheet0['Q34'] = '记录人'

    sheet0['B35'] = 'G1.1'
    sheet0['C35'] = 'G1.7'
    sheet0['D35'] = 'G2.3'
    sheet0['E35'] = 'G2.4'
    sheet0['F35'] = 'G5.1'
    sheet0['G35'] = 'G5.4'
    sheet0['H35'] = 'G5.7'
    sheet0['I35'] = 'G5.3'
    sheet0['J35'] = 'G5.8'
    sheet0['K35'] = 'G5.5'
    sheet0['L35'] = 'G5.2'
    sheet0['M35'] = 'G5.6'
    sheet0['N35'] = 'GE2.1'
    sheet0['O35'] = 'GE3.1'
    sheet0['P35'] = 'GE4.1'

    for i in range(36, 48):
        sheet0[f'A{i}'] = datetime.time(hour=(i - 16) % 24).strftime('%H:%M')

    # merge row18:row30 form A to Q
    for i in range(48, 58):
        sheet0.merge_cells(f'A{i}:Q{i}')

    sheet0.column_dimensions['A'].width = 11.67
    sheet0.column_dimensions['B'].width = 18.44
    sheet0.column_dimensions['C'].width = 13.78
    sheet0.column_dimensions['D'].width = 8.11
    sheet0.column_dimensions['E'].width = 8.11
    sheet0.column_dimensions['F'].width = 10.44
    sheet0.column_dimensions['G'].width = 15.67
    sheet0.column_dimensions['H'].width = 15.67
    sheet0.column_dimensions['I'].width = 15.67
    sheet0.column_dimensions['J'].width = 15.67
    sheet0.column_dimensions['K'].width = 8.11
    sheet0.column_dimensions['L'].width = 10.44
    sheet0.column_dimensions['M'].width = 17.22
    sheet0.column_dimensions['N'].width = 8.11
    sheet0.column_dimensions['O'].width = 8.11
    sheet0.column_dimensions['P'].width = 8.11
    sheet0.column_dimensions['Q'].width = 10.44

    for i in range(18, 31):
        sheet0.row_dimensions[i].height = 18
    for i in range(48, 58):
        sheet0.row_dimensions[i].height = 18

    font1 = Font(name="宋体", size=20, bold=True)
    sheet0['A1'].font = font1
    sheet0['A31'].font = font1

    font2 = Font(name="宋体", size=16, bold=True)
    for rowOfCells in sheet0['A3:D3']:
        for cell in rowOfCells:
            cell.font = font2
    # sheet0.row_dimensions[3].font = font2
    for rowOfCells in sheet0['A33:D33']:
        for cell in rowOfCells:
            cell.font = font2
    # sheet0.row_dimensions[4].font = font2
    for rowOfCells in sheet0['A4:Q4']:
        for cell in rowOfCells:
            cell.font = font2
    for rowOfCells in sheet0['A34:Q34']:
        for cell in rowOfCells:
            cell.font = font2

    font3 = Font(name="宋体", size=13, bold=True)

    sheet0.row_dimensions[5].font = font3
    for rowOfCells in sheet0['B5:P5']:
        for cell in rowOfCells:
            cell.font = font3
    for rowOfCells in sheet0['B35:P35']:
        for cell in rowOfCells:
            cell.font = font3
    for rowOfCells in sheet0['A6:A17']:
        for cell in rowOfCells:
            cell.font = font3
    for rowOfCells in sheet0['A36:A47']:
        for cell in rowOfCells:
            cell.font = font3

    font4 = Font(name="宋体", size=11, bold=False)
    for rowOfCells in sheet0['B6:P17']:
        for cell in rowOfCells:
            cell.font = font4
    for rowOfCells in sheet0['B36:P47']:
        for cell in rowOfCells:
            cell.font = font4

    align = Alignment(horizontal='center', vertical='center')
    sheet0['A1'].alignment = align
    sheet0['A31'].alignment = align

    for rowOfCells in sheet0['A3:Q17']:
        for cell in rowOfCells:
            cell.alignment = align
    for rowOfCells in sheet0['A33:Q47']:
        for cell in rowOfCells:
            cell.alignment = align

    # # 3. 烘烤温度

    # ## 3.1 设置固定值
    #

    # In[145]:

    sheet1.merge_cells('A1:O2')
    sheet1['A1'] = '2023年春季EAST真空烘烤温度记录表'
    sheet1['A3'] = '日期'
    sheet1['B3'] = '=真空运行!B3'
    sheet1['C3'] = '值班人员'
    sheet1.merge_cells('D3:Q3')
    sheet1['D3'] = '=真空运行!D3'

    # In[146]:

    sheet1.merge_cells('A4:A5')
    sheet1.merge_cells('C4:D4')
    sheet1.merge_cells('E4:I4')
    sheet1.merge_cells('M4:N4')
    sheet1.merge_cells('O4:O5')
    sheet1['A4'] = '时间'
    sheet1['B4'] = '硼化水管'
    sheet1['C4'] = '夹层'
    sheet1['E4'] = '第一壁'

    sheet1['J4'] = '上窗口'
    sheet1['K4'] = '水平窗'
    sheet1['L4'] = '下窗口'
    sheet1['M4'] = '抽气管道'
    sheet1['O4'] = '记录人'

    # In[147]:

    sheet1['B5'] = 'TVVG'
    sheet1['C5'] = 'TVVI'
    sheet1['D5'] = 'TVVM'
    sheet1['E5'] = 'THFM'
    sheet1['F5'] = 'TPG1'
    sheet1['G5'] = 'TPE18'
    sheet1['H5'] = 'TPO3'
    sheet1['I5'] = 'TPO15'
    sheet1['J5'] = 'TVUP'
    sheet1['K5'] = 'TVHF'
    sheet1['L5'] = 'TVLP'
    sheet1['M5'] = 'T2'
    sheet1['N5'] = 'T13'

    # In[148]:

    for i in range(6, 30):
        sheet1[f'A{i}'] = datetime.time(hour=(i + 2) % 24).strftime('%H:%M')

    # ## 3.2 设置行高和列高

    # In[149]:

    # sheet1.column_dimensions['A'].width = 11.67
    sheet1.column_dimensions['B'].width = 16.9
    sheet1.column_dimensions['C'].width = 12.5
    # sheet1.column_dimensions['D'].width = 8.11
    # sheet1.column_dimensions['E'].width = 8.11
    # sheet1.column_dimensions['F'].width = 10.44
    # sheet1.column_dimensions['G'].width = 15.67
    # sheet1.column_dimensions['H'].width = 15.67
    # sheet1.column_dimensions['I'].width = 15.67
    # sheet1.column_dimensions['J'].width = 15.67
    # sheet1.column_dimensions['K'].width = 8.11
    # sheet1.column_dimensions['L'].width = 10.44
    # sheet1.column_dimensions['M'].width = 17.22
    # sheet1.column_dimensions['N'].width = 8.11
    # sheet1.column_dimensions['O'].width = 8.11
    # sheet1.column_dimensions['P'].width = 8.11
    # sheet1.column_dimensions['Q'].width = 10.44

    # ## 3.3 设置样式

    # In[150]:

    sheet1['A1'].font = font1

    for rowOfCells in sheet1['A3:D3']:
        for cell in rowOfCells:
            cell.font = font2

    for rowOfCells in sheet1['A4:O4']:
        for cell in rowOfCells:
            cell.font = font2

    for rowOfCells in sheet1['B5:N5']:
        for cell in rowOfCells:
            cell.font = font3

    for rowOfCells in sheet1['A6:A29']:
        for cell in rowOfCells:
            cell.font = font3

    # ## 3.4 设置对齐

    # In[151]:

    for rowOfCells in sheet1['A1:O29']:
        for cell in rowOfCells:
            cell.alignment = align

    #

    # In[152]:

    date_today = (datetime.date.today() - datetime.timedelta(days=1)).strftime('%Y%m%d')
    file_name = f'{date_today}真空记录运行表.xlsx'
    wb.save(file_name)
    return file_name

